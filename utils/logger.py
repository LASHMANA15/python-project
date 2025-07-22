import csv
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook

log_csv = "logs/actions.csv"
log_excel = "logs/actions.xlsx"
os.makedirs("logs", exist_ok=True)

def log_action(name, dept, roll, action):
    time_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row = [name, dept, roll, action, time_now]

    # Write to CSV
    with open(log_csv, 'a', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(row)

    # Write to Excel
    if not os.path.exists(log_excel):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Department", "Roll Number", "Action", "Timestamp"])
        wb.save(log_excel)

    wb = load_workbook(log_excel)
    ws = wb.active
    ws.append(row)
    wb.save(log_excel)